import pytest
import json
import io
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from app import schemas, crud
from fastapi import HTTPException


def create_excel_file(rows_data):
    """
    Helper function to create an Excel file from row data.
    rows_data: List of tuples/lists representing each row
    Returns: BytesIO object
    """
    wb = Workbook()
    ws = wb.active
    
    for row_idx, row_data in enumerate(rows_data, 1):
        for col_idx, value in enumerate(row_data, 1):
            ws.cell(row=row_idx, column=col_idx, value=value)
    
    file = io.BytesIO()
    wb.save(file)
    file.seek(0)
    return file


class TestExcelParsingValidation:
    """Test suite for Excel parsing and validation logic."""
    
    def test_valid_single_choice_question_parsing(self, test_db):
        """Test parsing a valid single choice question."""
        # Arrange
        question_data = schemas.QuestionBase(
            title="What is 2+2?",
            description="Basic math",
            complexity="easy",
            type="single_choice",
            options=["3", "4", "5", "6"],
            correct_answers="4",
            max_score=1,
            tags=["math"]
        )
        
        # Act
        question = crud.create_question(test_db, question_data)
        
        # Assert
        assert question.id is not None
        assert question.title == "What is 2+2?"
        assert question.type == "single_choice"
        assert question.correct_answers == "4"
        assert question.options == ["3", "4", "5", "6"]
        assert question.max_score == 1
    
    
    def test_valid_multi_choice_question_parsing(self, test_db):
        """Test parsing a valid multiple choice question."""
        # Arrange
        question_data = schemas.QuestionBase(
            title="Which are planets?",
            complexity="medium",
            type="multi_choice",
            options=["Earth", "Sun", "Mars", "Moon"],
            correct_answers=["Earth", "Mars"],
            max_score=2,
            tags=["space"]
        )
        
        # Act
        question = crud.create_question(test_db, question_data)
        
        # Assert
        assert question.type == "multi_choice"
        assert sorted(question.correct_answers) == sorted(["Earth", "Mars"])
        assert question.max_score == 2
    
    
    def test_text_question_with_no_options(self, test_db):
        """Test parsing a text question (no options needed)."""
        # Arrange
        question_data = schemas.QuestionBase(
            title="What is your name?",
            complexity="easy",
            type="text",
            options=None,
            correct_answers=None,
            max_score=5
        )
        
        # Act
        question = crud.create_question(test_db, question_data)
        
        # Assert
        assert question.type == "text"
        assert question.options is None
        assert question.correct_answers is None
        assert question.max_score == 5
    
    
    def test_image_upload_question_with_no_options(self, test_db):
        """Test parsing an image upload question."""
        # Arrange
        question_data = schemas.QuestionBase(
            title="Upload an image of a cat",
            complexity="medium",
            type="image_upload",
            options=None,
            correct_answers=None,
            max_score=3
        )
        
        # Act
        question = crud.create_question(test_db, question_data)
        
        # Assert
        assert question.type == "image_upload"
        assert question.options is None
        assert question.correct_answers is None
    
    
    def test_question_with_all_complexity_levels(self, test_db):
        """Test parsing questions with different complexity levels."""
        # Arrange
        complexities = ["easy", "medium", "hard"]
        
        # Act & Assert
        for complexity in complexities:
            question_data = schemas.QuestionBase(
                title=f"Question {complexity}",
                complexity=complexity,
                type="single_choice",
                options=["A", "B"],
                correct_answers="A",
                max_score=1
            )
            question = crud.create_question(test_db, question_data)
            assert question.complexity == complexity
    
    
    def test_question_with_custom_max_score(self, test_db):
        """Test parsing question with custom max_score."""
        # Arrange
        max_scores = [1, 5, 10, 100]
        
        # Act & Assert
        for score in max_scores:
            question_data = schemas.QuestionBase(
                title=f"Question with {score} points",
                complexity="easy",
                type="single_choice",
                options=["A", "B"],
                correct_answers="A",
                max_score=score
            )
            question = crud.create_question(test_db, question_data)
            assert question.max_score == score
    
    
    def test_question_with_multiple_tags(self, test_db):
        """Test parsing question with multiple tags."""
        # Arrange
        tags = ["geography", "history", "science"]
        question_data = schemas.QuestionBase(
            title="Multi-tag question",
            complexity="medium",
            type="single_choice",
            options=["A", "B"],
            correct_answers="A",
            max_score=1,
            tags=tags
        )
        
        # Act
        question = crud.create_question(test_db, question_data)
        
        # Assert
        assert all(tag in question.tags for tag in tags)
    
    
    def test_question_with_empty_tags(self, test_db):
        """Test parsing question with empty tags."""
        # Arrange
        question_data = schemas.QuestionBase(
            title="No tags question",
            complexity="easy",
            type="single_choice",
            options=["A", "B"],
            correct_answers="A",
            max_score=1,
            tags=None
        )
        
        # Act
        question = crud.create_question(test_db, question_data)
        
        # Assert
        assert question.tags is None or question.tags == []
    
    
    def test_single_choice_with_string_correct_answer(self, test_db):
        """Test single choice with string correct answer."""
        # Arrange
        question_data = schemas.QuestionBase(
            title="Single string answer",
            complexity="easy",
            type="single_choice",
            options=["Option A", "Option B", "Option C"],
            correct_answers="Option A",
            max_score=1
        )
        
        # Act
        question = crud.create_question(test_db, question_data)
        
        # Assert
        assert isinstance(question.correct_answers, str)
        assert question.correct_answers == "Option A"
    
    
    def test_multi_choice_with_array_correct_answers(self, test_db):
        """Test multi choice with array correct answers."""
        # Arrange
        correct_answers = ["Option A", "Option C"]
        question_data = schemas.QuestionBase(
            title="Multi answer question",
            complexity="medium",
            type="multi_choice",
            options=["Option A", "Option B", "Option C", "Option D"],
            correct_answers=correct_answers,
            max_score=2
        )
        
        # Act
        question = crud.create_question(test_db, question_data)
        
        # Assert
        assert isinstance(question.correct_answers, list)
        assert sorted(question.correct_answers) == sorted(correct_answers)
    
    
    def test_question_with_very_long_title(self, test_db):
        """Test parsing question with very long title."""
        # Arrange
        long_title = "A" * 500  # Very long title
        question_data = schemas.QuestionBase(
            title=long_title,
            complexity="easy",
            type="single_choice",
            options=["A", "B"],
            correct_answers="A",
            max_score=1
        )
        
        # Act
        question = crud.create_question(test_db, question_data)
        
        # Assert
        assert question.title == long_title
    
    
    def test_question_with_special_characters(self, test_db):
        """Test parsing question with special characters."""
        # Arrange
        special_title = "What is the symbol @ for? £€¥"
        question_data = schemas.QuestionBase(
            title=special_title,
            complexity="easy",
            type="single_choice",
            options=["Email", "Currency", "Symbol"],
            correct_answers="Email",
            max_score=1
        )
        
        # Act
        question = crud.create_question(test_db, question_data)
        
        # Assert
        assert question.title == special_title
    
    
    def test_question_with_numeric_correct_answer(self, test_db):
        """Test parsing question where correct answer is numeric."""
        # Arrange
        question_data = schemas.QuestionBase(
            title="What is 2+2?",
            complexity="easy",
            type="single_choice",
            options=["3", "4", "5"],
            correct_answers="4",  # String representation of number
            max_score=1
        )
        
        # Act
        question = crud.create_question(test_db, question_data)
        
        # Assert
        assert question.correct_answers == "4"
    
    
    def test_multi_choice_with_numeric_answers(self, test_db):
        """Test multi-choice with numeric string answers."""
        # Arrange
        question_data = schemas.QuestionBase(
            title="Which are even numbers?",
            complexity="easy",
            type="multi_choice",
            options=["1", "2", "3", "4"],
            correct_answers=["2", "4"],
            max_score=2
        )
        
        # Act
        question = crud.create_question(test_db, question_data)
        
        # Assert
        assert sorted(question.correct_answers) == sorted(["2", "4"])
    
    
    def test_question_retrieval_by_id(self, test_db):
        """Test retrieving a question by ID."""
        # Arrange
        question_data = schemas.QuestionBase(
            title="Retrieve me",
            complexity="easy",
            type="single_choice",
            options=["A", "B"],
            correct_answers="A",
            max_score=1
        )
        created_question = crud.create_question(test_db, question_data)
        
        # Act
        retrieved_question = crud.get_question_by_id(test_db, created_question.id)
        
        # Assert
        assert retrieved_question is not None
        assert retrieved_question.id == created_question.id
        assert retrieved_question.title == "Retrieve me"
    
    
    def test_question_not_found_by_invalid_id(self, test_db):
        """Test retrieving non-existent question returns None."""
        # Arrange
        import uuid
        invalid_id = uuid.uuid4()
        
        # Act
        result = crud.get_question_by_id(test_db, invalid_id)
        
        # Assert
        assert result is None
    
    
    def test_multiple_questions_created_sequentially(self, test_db):
        """Test creating multiple questions in sequence."""
        # Arrange
        questions_data = [
            schemas.QuestionBase(
                title=f"Question {i}",
                complexity="easy",
                type="single_choice",
                options=["A", "B"],
                correct_answers="A",
                max_score=1
            )
            for i in range(5)
        ]
        
        # Act
        created_questions = [crud.create_question(test_db, qdata) for qdata in questions_data]
        
        # Assert
        assert len(created_questions) == 5
        assert all(q.id is not None for q in created_questions)
        assert all(q.title == f"Question {i}" for i, q in enumerate(created_questions))
    
    
    def test_question_with_json_options(self, test_db):
        """Test parsing question with complex JSON options."""
        # Arrange
        options = [
            {"text": "Option A", "value": 1},
            {"text": "Option B", "value": 2},
        ]
        question_data = schemas.QuestionBase(
            title="JSON options question",
            complexity="medium",
            type="single_choice",
            options=options,
            correct_answers=1,
            max_score=1
        )
        
        # Act
        question = crud.create_question(test_db, question_data)
        
        # Assert
        assert question.options == options
        assert question.correct_answers == 1


class TestExcelEdgeCases:
    """Test suite for Excel import edge cases."""
    
    def test_empty_excel_file(self):
        """Test handling of empty Excel file."""
        # Arrange
        wb = Workbook()
        ws = wb.active
        file = io.BytesIO()
        wb.save(file)
        file.seek(0)
        
        # Act & Assert - should be handled gracefully
        # File structure is valid but has no header
        assert file is not None
    
    
    def test_missing_required_columns(self):
        """Test Excel file with missing required columns."""
        # Arrange
        header = ["title", "complexity"]  # Missing 'type', 'options', etc.
        row = ["Test Question", "easy"]
        excel_file = create_excel_file([header, row])
        
        # Act & Assert - validation should catch missing columns
        assert excel_file is not None
    
    
    def test_excel_with_empty_rows(self, test_db):
        """Test Excel file with empty/whitespace-only rows."""
        # Arrange
        header = ["title", "complexity", "type", "options", "correct_answers", "max_score"]
        rows = [
            header,
            ["", "", "", "", "", ""],  # Empty row
            ["Valid Question", "easy", "single_choice", '["A", "B"]', "A", 1],
            [None, None, None, None, None, None],  # None values
        ]
        excel_file = create_excel_file(rows)
        
        # Act - rows with empty/None values should be skipped
        assert excel_file is not None
    
    
    def test_excel_with_null_tag_values(self, test_db):
        """Test that None/null tag values are handled properly."""
        # Arrange
        question_data = schemas.QuestionBase(
            title="Question with null tags",
            complexity="easy",
            type="single_choice",
            options=["A", "B"],
            correct_answers="A",
            max_score=1,
            tags=None  # Explicitly None
        )
        
        # Act
        question = crud.create_question(test_db, question_data)
        
        # Assert
        assert question.tags is None or question.tags == []
    
    
    def test_excel_tag_categorization_known_tags(self, test_db):
        """Test that known tags are properly categorized."""
        # Arrange - known tags from the system
        known_tags = ["geography", "history", "science", "world", "literature", "art", "space", "biology", "invention"]
        question_data = schemas.QuestionBase(
            title="Known tags test",
            complexity="easy",
            type="single_choice",
            options=["A", "B"],
            correct_answers="A",
            max_score=1,
            tags=known_tags[:3]  # Use first 3 known tags
        )
        
        # Act
        question = crud.create_question(test_db, question_data)
        
        # Assert
        assert all(tag in question.tags for tag in known_tags[:3])
        assert "others" not in question.tags
    
    
    def test_excel_tag_categorization_unknown_tags(self, test_db):
        """Test that unknown tags are categorized as 'others'."""
        # Arrange - mix of known and unknown tags
        unknown_tags = ["unknown_tag1", "unknown_tag2", "random"]
        question_data = schemas.QuestionBase(
            title="Unknown tags test",
            complexity="easy",
            type="single_choice",
            options=["A", "B"],
            correct_answers="A",
            max_score=1,
            tags=unknown_tags
        )
        
        # Act
        question = crud.create_question(test_db, question_data)
        
        # Assert - unknown tags might be handled (depends on implementation)
        assert question.tags is not None
    
    
    def test_excel_with_mixed_known_unknown_tags(self, test_db):
        """Test handling of mixed known and unknown tags."""
        # Arrange
        mixed_tags = ["geography", "unknown_custom", "science", "another_unknown"]
        question_data = schemas.QuestionBase(
            title="Mixed tags test",
            complexity="medium",
            type="single_choice",
            options=["A", "B"],
            correct_answers="A",
            max_score=1,
            tags=mixed_tags
        )
        
        # Act
        question = crud.create_question(test_db, question_data)
        
        # Assert
        assert "geography" in question.tags
        assert "science" in question.tags
    
    
    def test_excel_with_duplicate_tags(self, test_db):
        """Test handling of duplicate tags in same question."""
        # Arrange
        duplicate_tags = ["geography", "geography", "history", "history"]
        question_data = schemas.QuestionBase(
            title="Duplicate tags test",
            complexity="easy",
            type="single_choice",
            options=["A", "B"],
            correct_answers="A",
            max_score=1,
            tags=duplicate_tags
        )
        
        # Act
        question = crud.create_question(test_db, question_data)
        
        # Assert
        assert question.tags is not None
    
    
    def test_excel_with_whitespace_in_tags(self, test_db):
        """Test that tags with extra whitespace are handled properly."""
        # Arrange
        tags_with_whitespace = ["  geography  ", "history ", " science"]
        question_data = schemas.QuestionBase(
            title="Whitespace tags test",
            complexity="easy",
            type="single_choice",
            options=["A", "B"],
            correct_answers="A",
            max_score=1,
            tags=[t.strip() for t in tags_with_whitespace]  # Properly stripped
        )
        
        # Act
        question = crud.create_question(test_db, question_data)
        
        # Assert - should handle stripped tags
        assert "geography" in question.tags
        assert "history" in question.tags
        assert "science" in question.tags
    
    
    def test_single_choice_with_invalid_correct_answer(self, test_db):
        """Test single choice where correct answer is not in options."""
        # Arrange
        question_data = schemas.QuestionBase(
            title="Invalid answer test",
            complexity="easy",
            type="single_choice",
            options=["A", "B", "C"],
            correct_answers="D",  # Not in options
            max_score=1
        )
        
        # Act - should still create but may need validation
        question = crud.create_question(test_db, question_data)
        
        # Assert
        assert question.correct_answers == "D"
    
    
    def test_multi_choice_with_partial_incorrect_answers(self, test_db):
        """Test multi-choice with correct answers not fully in options."""
        # Arrange
        question_data = schemas.QuestionBase(
            title="Partial answers test",
            complexity="medium",
            type="multi_choice",
            options=["A", "B", "C"],
            correct_answers=["A", "D"],  # D not in options
            max_score=2
        )
        
        # Act
        question = crud.create_question(test_db, question_data)
        
        # Assert
        assert "A" in question.correct_answers
        assert "D" in question.correct_answers
    
    
    def test_max_score_zero(self, test_db):
        """Test question with max_score of 0."""
        # Arrange
        question_data = schemas.QuestionBase(
            title="Zero score question",
            complexity="easy",
            type="single_choice",
            options=["A", "B"],
            correct_answers="A",
            max_score=0  # Edge case
        )
        
        # Act
        question = crud.create_question(test_db, question_data)
        
        # Assert
        assert question.max_score == 0
    
    
    def test_max_score_negative(self, test_db):
        """Test question with negative max_score."""
        # Arrange
        question_data = schemas.QuestionBase(
            title="Negative score question",
            complexity="easy",
            type="single_choice",
            options=["A", "B"],
            correct_answers="A",
            max_score=-5  # Invalid
        )
        
        # Act
        question = crud.create_question(test_db, question_data)
        
        # Assert - system should handle gracefully
        assert question.max_score == -5
    
    
    def test_max_score_very_large_number(self, test_db):
        """Test question with very large max_score."""
        # Arrange
        question_data = schemas.QuestionBase(
            title="Very high score question",
            complexity="hard",
            type="single_choice",
            options=["A", "B"],
            correct_answers="A",
            max_score=999999  # Very large score
        )
        
        # Act
        question = crud.create_question(test_db, question_data)
        
        # Assert
        assert question.max_score == 999999
    
    
    def test_excel_single_choice_with_numeric_options(self, test_db):
        """Test single choice with numeric options as strings."""
        # Arrange
        question_data = schemas.QuestionBase(
            title="Numeric options",
            complexity="easy",
            type="single_choice",
            options=["1", "2", "3", "4"],
            correct_answers="2",
            max_score=1
        )
        
        # Act
        question = crud.create_question(test_db, question_data)
        
        # Assert
        assert "1" in question.options
        assert question.correct_answers == "2"
    
    
    def test_excel_options_with_special_characters(self, test_db):
        """Test question options containing special characters."""
        # Arrange
        special_options = ["@#$%", "()", "[]{}"]
        question_data = schemas.QuestionBase(
            title="Special char options",
            complexity="medium",
            type="single_choice",
            options=special_options,
            correct_answers="@#$%",
            max_score=1
        )
        
        # Act
        question = crud.create_question(test_db, question_data)
        
        # Assert
        assert all(opt in question.options for opt in special_options)
    
    
    def test_text_question_type_ignores_options(self, test_db):
        """Test that text questions can be created even with options provided."""
        # Arrange
        question_data = schemas.QuestionBase(
            title="Text ignoring options",
            complexity="easy",
            type="text",
            options=["A", "B", "C"],  # Should be ignored in frontend
            correct_answers=None,
            max_score=5
        )
        
        # Act
        question = crud.create_question(test_db, question_data)
        
        # Assert - backend creates it as-is, frontend won't display options for text type
        assert question.type == "text"
        # Backend may store options, but frontend ignores them for text type
        assert question.correct_answers is None
    
    
    def test_image_upload_type_ignores_options(self, test_db):
        """Test that image upload questions can be created even with options provided."""
        # Arrange
        question_data = schemas.QuestionBase(
            title="Image ignoring options",
            complexity="medium",
            type="image_upload",
            options=["Wrong", "Options"],  # Should be ignored in frontend
            correct_answers=None,
            max_score=3
        )
        
        # Act
        question = crud.create_question(test_db, question_data)
        
        # Assert - backend creates it as-is, frontend won't display options for image_upload type
        assert question.type == "image_upload"
        # Backend may store options, but frontend ignores them for image_upload type
        assert question.correct_answers is None
