import json
from typing import Any
from uuid import UUID
from datetime import datetime, timezone

from fastapi import APIRouter, Depends, File, UploadFile, HTTPException, status, Query, Body
from fastapi.responses import JSONResponse, StreamingResponse
from sqlalchemy.orm import Session
import openpyxl

from ..database import get_db
from .. import schemas, crud, models, security

router = APIRouter(prefix="/admin", tags=["Admin"])


def get_current_admin_user(
    current_user: models.User = Depends(security.get_current_user),
) -> models.User:
    if current_user.role != "admin":
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="Admin privileges required",
        )
    return current_user


@router.get("/questions/", response_model=list[schemas.Question])
def list_questions(
    db: Session = Depends(get_db),
    _: models.User = Depends(get_current_admin_user),
    search: str = Query(None, description="Search in title and complexity"),
    question_type: str = Query(None, description="Filter by question type"),
    complexity: str = Query(None, description="Filter by complexity"),
    tag: str = Query(None, description="Filter by tag"),
):
    """List all questions for exam building with search and filter capabilities."""
    query = db.query(models.Question)
    
    if search:
        search_pattern = f"%{search}%"
        query = query.filter(
            models.Question.title.ilike(search_pattern) | 
            models.Question.complexity.ilike(search_pattern)
        )
    
    if question_type:
        query = query.filter(models.Question.type == question_type)
    
    if complexity:
        # Normalize complexity to lowercase
        complexity = complexity.lower()
        query = query.filter(models.Question.complexity == complexity)
    
    if tag:
        # Filter by tag - since tags is a JSON array, we need to check if tag is in the array
        # Using PostgreSQL's JSON containment operator
        from sqlalchemy import and_, or_
        tag = tag.lower()
        # Check if the tag is in the JSON array or if "others" is requested and no exact match
        query = query.filter(
            models.Question.tags.astext.contains(f'"{tag}"')
        )
    
    return query.all()


@router.get("/questions/{question_id}", response_model=schemas.Question)
def get_question(
    question_id: UUID,
    db: Session = Depends(get_db),
    _: models.User = Depends(get_current_admin_user),
):
    """Get a single question by ID."""
    question = crud.get_question_by_id(db, question_id)
    if not question:
        raise HTTPException(status_code=404, detail="Question not found")
    return question


@router.delete("/questions/{question_id}")
def delete_question(
    question_id: UUID,
    db: Session = Depends(get_db),
    _: models.User = Depends(get_current_admin_user),
):
    """Delete a question by ID."""
    question = crud.get_question_by_id(db, question_id)
    if not question:
        raise HTTPException(status_code=404, detail="Question not found")
    
    db.delete(question)
    db.commit()
    return {"status": "success", "message": "Question deleted successfully"}


@router.post("/questions/delete-bulk")
def delete_bulk_questions(
    request_body: dict = Body(...),
    db: Session = Depends(get_db),
    _: models.User = Depends(get_current_admin_user),
):
    """Delete multiple questions by IDs."""
    print(f"DEBUG: Received request body: {request_body}")
    question_ids = request_body.get("question_ids", [])
    print(f"DEBUG: Question IDs to delete: {question_ids}")
    deleted_count = 0
    failed_count = 0
    
    for question_id in question_ids:
        try:
            qid = UUID(str(question_id))
            print(f"DEBUG: Processing question {qid}")
            
            # First, find all answers for this question
            answers = db.query(models.Answer).filter(
                models.Answer.question_id == qid
            ).all()
            print(f"DEBUG: Found {len(answers)} answers for question {qid}")
            
            # Delete evaluations that reference these answers
            for answer in answers:
                db.query(models.Evaluation).filter(
                    models.Evaluation.answer_id == answer.id
                ).delete(synchronize_session=False)
            
            # Then delete all answers for this question
            db.query(models.Answer).filter(
                models.Answer.question_id == qid
            ).delete(synchronize_session=False)
            
            # Finally delete the question
            question = crud.get_question_by_id(db, qid)
            print(f"DEBUG: Found question {qid}: {question is not None}")
            if question:
                db.delete(question)
                deleted_count += 1
            
            # Commit after each successful deletion to prevent transaction abort
            db.commit()
            print(f"DEBUG: Successfully deleted question {qid}")
            
        except Exception as e:
            print(f"DEBUG: Error deleting question {question_id}: {e}")
            db.rollback()  # Rollback on error
            failed_count += 1
    
    print(f"DEBUG: Deleted {deleted_count} questions, {failed_count} failed")
    return {
        "status": "success",
        "deleted": deleted_count,
        "failed": failed_count
    }


@router.post("/upload-questions/")
async def upload_questions(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    _: models.User = Depends(get_current_admin_user),
):
    # Validate file extension
    if not file.filename:
        raise HTTPException(status_code=400, detail="File name is required")
    
    if not file.filename.lower().endswith(('.xlsx', '.xls')):
        raise HTTPException(
            status_code=400, 
            detail="Only Excel files (.xlsx or .xls) are accepted. Please upload a valid Excel file."
        )
    
    try:
        wb = openpyxl.load_workbook(file.file, data_only=True)
        ws = wb.active
    except Exception as e:
        print(f"Excel load error: {e}")
        raise HTTPException(status_code=400, detail=f"Invalid Excel file: {e}")

    # Validate header columns and build index map
    try:
        header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
        header = [str(h).strip().lower() if h is not None else "" for h in header_row]
    except StopIteration:
        raise HTTPException(status_code=400, detail="Excel file is empty")

    expected = [
        "title",
        "complexity",
        "type",
        "options",
        "correct_answers",
        "max_score",
    ]
    missing = [c for c in expected if c not in header]
    if missing:
        raise HTTPException(
            status_code=400,
            detail=f"Missing required columns: {', '.join(missing)}",
        )
    idx = {name: header.index(name) for name in expected}
    
    # Check if optional columns exist
    tags_idx = header.index("tags") if "tags" in header else None
    desc_idx = header.index("description") if "description" in header else None

    allowed_types = {"single_choice", "multi_choice", "text", "image_upload"}
    allowed_complexity = {"easy", "medium", "hard"}
    known_tags = {"geography", "history", "science", "world", "literature", "art", "space", "biology", "invention"}

    imported = 0
    for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not row:
            continue
        # Extract using header indices for flexibility
        title = row[idx["title"]] if idx["title"] < len(row) else None
        complexity = row[idx["complexity"]] if idx["complexity"] < len(row) else None
        qtype = row[idx["type"]] if idx["type"] < len(row) else None
        options_raw = row[idx["options"]] if idx["options"] < len(row) else None
        correct_raw = row[idx["correct_answers"]] if idx["correct_answers"] < len(row) else None
        max_score_raw = row[idx["max_score"]] if idx["max_score"] < len(row) else None
        tags_raw = row[tags_idx] if tags_idx is not None and tags_idx < len(row) else None
        description = row[desc_idx] if desc_idx is not None and desc_idx < len(row) else None

        if not title or not complexity or not qtype:
            # Skip incomplete rows
            continue

        # Normalize complexity to lowercase
        complexity = str(complexity).strip().lower()
        if complexity not in allowed_complexity:
            raise HTTPException(
                status_code=400,
                detail=f"Row {i}: invalid complexity '{complexity}'. Allowed: {sorted(allowed_complexity)}",
            )

        qtype = str(qtype).strip()
        if qtype not in allowed_types:
            raise HTTPException(
                status_code=400,
                detail=f"Row {i}: invalid type '{qtype}'. Allowed: {sorted(allowed_types)}",
            )

        try:
            # Parse tags
            tags = []
            if tags_raw:
                if isinstance(tags_raw, str):
                    # Split by comma and strip whitespace
                    tags_list = [t.strip().lower() for t in tags_raw.split(",") if t.strip()]
                elif isinstance(tags_raw, list):
                    tags_list = [t.lower() for t in tags_raw if t]
                else:
                    tags_list = []
                
                # Categorize tags: known tags stay as-is, unknown tags become "others"
                for tag in tags_list:
                    if tag in known_tags:
                        tags.append(tag)
                    else:
                        if "others" not in tags:
                            tags.append("others")
            
            # Parse options
            options: Any = None
            if isinstance(options_raw, str):
                if options_raw.strip() not in ("", "null"):
                    options = json.loads(options_raw)
            elif options_raw not in (None, ""):
                options = options_raw

            # Parse correct answers
            if isinstance(correct_raw, str):
                correct_answers: Any = json.loads(correct_raw)
            else:
                correct_answers = correct_raw

            # Validate by type
            if qtype == "text" or qtype == "image_upload":
                # For text and image_upload questions, options and correct_answers should be null
                options = None
                correct_answers = None
            elif qtype == "single_choice":
                # options should be an array; correct_answers should be a single value
                if options is None or not isinstance(options, list):
                    raise ValueError("'options' must be a JSON array for single_choice")
                if isinstance(correct_answers, list):
                    raise ValueError("'correct_answers' must be a single JSON value for single_choice")
            elif qtype == "multi_choice":
                # options should be an array; correct_answers should be an array
                if options is None or not isinstance(options, list):
                    raise ValueError("'options' must be a JSON array for multi_choice")
                if correct_answers is None or not isinstance(correct_answers, list):
                    raise ValueError("'correct_answers' must be a JSON array for multi_choice")

            # For non-text types, correct_answers is required
            if qtype in {"single_choice", "multi_choice"} and correct_answers in (None, ""):
                raise ValueError("'correct_answers' is required for single_choice and multi_choice")

            # max score
            max_score_val = int(max_score_raw) if max_score_raw not in (None, "") else 1
            if max_score_val <= 0:
                raise ValueError("max_score must be positive")
        except Exception as e:
            raise HTTPException(status_code=400, detail=f"Row {i} parse error: {e}")

        q_schema = schemas.QuestionBase(
            title=str(title),
            description=str(description) if description else None,
            complexity=complexity,
            type=qtype,
            options=options,
            correct_answers=correct_answers,
            max_score=max_score_val,
            tags=tags if tags else None,
        )
        crud.create_question(db, q_schema)
        imported += 1

    return JSONResponse({"status": "success", "rows_imported": imported})


@router.post("/preview-questions/")
async def preview_questions(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    _: models.User = Depends(get_current_admin_user),
):
    """Preview questions from Excel file without importing them."""
    # Validate file extension
    if not file.filename:
        raise HTTPException(status_code=400, detail="File name is required")
    
    if not file.filename.lower().endswith(('.xlsx', '.xls')):
        raise HTTPException(
            status_code=400, 
            detail="Only Excel files (.xlsx or .xls) are accepted. Please upload a valid Excel file."
        )
    
    try:
        wb = openpyxl.load_workbook(file.file, data_only=True)
        ws = wb.active
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Invalid Excel file: {e}")

    # Validate header columns and build index map
    try:
        header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
        header = [str(h).strip().lower() if h is not None else "" for h in header_row]
    except StopIteration:
        raise HTTPException(status_code=400, detail="Excel file is empty")

    expected = [
        "title",
        "complexity",
        "type",
        "options",
        "correct_answers",
        "max_score",
    ]
    missing = [c for c in expected if c not in header]
    if missing:
        raise HTTPException(
            status_code=400,
            detail=f"Missing required columns: {', '.join(missing)}",
        )
    idx = {name: header.index(name) for name in expected}
    
    # Check if optional columns exist
    tags_idx = header.index("tags") if "tags" in header else None
    desc_idx = header.index("description") if "description" in header else None

    allowed_types = {"single_choice", "multi_choice", "text", "image_upload"}
    preview_data = []
    for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not row:
            continue
        # Extract using header indices for flexibility
        title = row[idx["title"]] if idx["title"] < len(row) else None
        complexity = row[idx["complexity"]] if idx["complexity"] < len(row) else None
        qtype = row[idx["type"]] if idx["type"] < len(row) else None
        options_raw = row[idx["options"]] if idx["options"] < len(row) else None
        correct_raw = row[idx["correct_answers"]] if idx["correct_answers"] < len(row) else None
        max_score_raw = row[idx["max_score"]] if idx["max_score"] < len(row) else None

        if not title or not complexity or not qtype:
            # Skip incomplete rows
            continue

        qtype = str(qtype).strip()
        if qtype not in allowed_types:
            raise HTTPException(
                status_code=400,
                detail=f"Row {i}: invalid type '{qtype}'. Allowed: {sorted(allowed_types)}",
            )

        try:
            # Parse options
            options: Any = None
            if isinstance(options_raw, str):
                if options_raw.strip() not in ("", "null"):
                    options = json.loads(options_raw)
            elif options_raw not in (None, ""):
                options = options_raw

            # Parse correct answers
            if isinstance(correct_raw, str):
                correct_answers: Any = json.loads(correct_raw)
            else:
                correct_answers = correct_raw

            # Validate by type
            if qtype == "text" or qtype == "image_upload":
                # For text and image_upload questions, options and correct_answers should be null
                options = None
                correct_answers = None
            elif qtype == "single_choice":
                # options should be an array; correct_answers should be a single value
                if options is None or not isinstance(options, list):
                    raise ValueError("'options' must be a JSON array for single_choice")
                if isinstance(correct_answers, list):
                    raise ValueError("'correct_answers' must be a single JSON value for single_choice")
            elif qtype == "multi_choice":
                # options should be an array; correct_answers should be an array
                if options is None or not isinstance(options, list):
                    raise ValueError("'options' must be a JSON array for multi_choice")
                if correct_answers is None or not isinstance(correct_answers, list):
                    raise ValueError("'correct_answers' must be a JSON array for multi_choice")

            # For non-text types, correct_answers is required
            if qtype in {"single_choice", "multi_choice"} and correct_answers in (None, ""):
                raise ValueError("'correct_answers' is required for single_choice and multi_choice")

            # max score
            max_score_val = int(max_score_raw) if max_score_raw not in (None, "") else 1
            if max_score_val <= 0:
                raise ValueError("max_score must be positive")
        except Exception as e:
            raise HTTPException(status_code=400, detail=f"Row {i} parse error: {e}")

        preview_data.append({
            "title": str(title),
            "complexity": str(complexity),
            "type": qtype,
            "options": options,
            "correct_answers": correct_answers,
            "max_score": max_score_val,
        })

    return JSONResponse({"questions": preview_data, "count": len(preview_data)})


@router.get("/questions-template")
def download_questions_template(_: models.User = Depends(get_current_admin_user)):
    """Return an Excel template with headers and sample rows."""
    from openpyxl import Workbook
    from io import BytesIO

    wb = Workbook()
    ws = wb.active
    ws.title = "questions"
    headers = [
        "title",
        "complexity",
        "type",
        "options",
        "correct_answers",
        "max_score",
        "tags",
    ]
    ws.append(headers)
    # Sample rows
    ws.append([
        "What is 2+2?",
        "easy",
        "single_choice",
        "[\"1\", \"2\", \"3\", \"4\"]",
        "[\"4\"]",
        1,
        "math,arithmetic",
    ])
    ws.append([
        "Select prime numbers",
        "medium",
        "multi_choice",
        "[\"2\", \"4\", \"5\", \"6\"]",
        "[\"2\", \"5\"]",
        2,
        "math",
    ])
    ws.append([
        "Briefly explain polymorphism",
        "hard",
        "text",
        "null",
        "null",
        5,
        "programming,oop",
    ])
    ws.append([
        "Upload a diagram of the water cycle",
        "medium",
        "image_upload",
        "null",
        "null",
        10,
        "science,biology",
    ])

    stream = BytesIO()
    wb.save(stream)
    stream.seek(0)
    return StreamingResponse(
        stream,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": "attachment; filename=questions_template.xlsx"
        },
    )


@router.post("/exams/", response_model=schemas.Exam)
def create_exam(
    exam: schemas.ExamCreate,
    db: Session = Depends(get_db),
    _: models.User = Depends(get_current_admin_user),
):
    """Create a new exam with associated questions."""
    # Validate that all question IDs exist
    for question_id in exam.question_ids:
        question = crud.get_question_by_id(db, question_id)
        if not question:
            raise HTTPException(
                status_code=404,
                detail=f"Question with ID {question_id} not found",
            )
    
    return crud.create_exam(db, exam)


@router.get("/exams/")
def list_exams(
    db: Session = Depends(get_db),
    _: models.User = Depends(get_current_admin_user),
):
    """List all exams with publisher information."""
    exams = crud.get_exams(db)
    result = []
    for exam in exams:
        # Get publisher email if published
        publisher_email = None
        if exam.published_by:
            publisher = db.query(models.User).filter(models.User.id == exam.published_by).first()
            publisher_email = publisher.email if publisher else "Unknown"
        
        # Manually build exam dict to avoid Pydantic validation issues
        exam_dict = {
            "id": str(exam.id),
            "title": exam.title,
            "start_time": exam.start_time.isoformat(),
            "end_time": exam.end_time.isoformat(),
            "duration_minutes": exam.duration_minutes,
            "is_published": exam.is_published,
            "published_by": publisher_email,
            "target_candidates": exam.target_candidates,
            "questions": [
                {
                    "id": str(q.id),
                    "title": q.title,
                    "complexity": q.complexity,
                    "type": q.type,
                    "options": q.options,
                    "correct_answers": q.correct_answers,
                    "max_score": q.max_score,
                    "tags": q.tags,
                }
                for q in (exam.questions or [])
            ],
        }
        result.append(exam_dict)
    return result


@router.post("/exams/{exam_id}/publish")
def publish_exam(
    exam_id: UUID,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_admin_user),
):
    """Publish an exam to make it available to students."""
    exam = crud.get_exam_by_id(db, exam_id)
    if not exam:
        raise HTTPException(status_code=404, detail="Exam not found")
    
    exam.is_published = True
    exam.published_by = current_user.id  # Track which admin published
    db.commit()
    db.refresh(exam)
    
    # Get publisher email
    publisher_email = current_user.email
    
    # Manually build response
    return {
        "id": str(exam.id),
        "title": exam.title,
        "start_time": exam.start_time.isoformat(),
        "end_time": exam.end_time.isoformat(),
        "duration_minutes": exam.duration_minutes,
        "is_published": exam.is_published,
        "published_by": publisher_email,
        "target_candidates": exam.target_candidates,
        "questions": [
            {
                "id": str(q.id),
                "title": q.title,
                "complexity": q.complexity,
                "type": q.type,
                "options": q.options,
                "correct_answers": q.correct_answers,
                "max_score": q.max_score,
                "tags": q.tags,
            }
            for q in (exam.questions or [])
        ],
    }


@router.post("/exams/{exam_id}/unpublish")
def unpublish_exam(
    exam_id: UUID,
    db: Session = Depends(get_db),
    _: models.User = Depends(get_current_admin_user),
):
    """Unpublish an exam to make it unavailable to students."""
    exam = crud.get_exam_by_id(db, exam_id)
    if not exam:
        raise HTTPException(status_code=404, detail="Exam not found")
    
    exam.is_published = False
    db.commit()
    db.refresh(exam)
    
    # Get publisher email if exists
    publisher_email = None
    if exam.published_by:
        publisher = db.query(models.User).filter(models.User.id == exam.published_by).first()
        publisher_email = publisher.email if publisher else "Unknown"
    
    # Manually build response
    return {
        "id": str(exam.id),
        "title": exam.title,
        "start_time": exam.start_time.isoformat(),
        "end_time": exam.end_time.isoformat(),
        "duration_minutes": exam.duration_minutes,
        "is_published": exam.is_published,
        "published_by": publisher_email,
        "target_candidates": exam.target_candidates,
        "questions": [
            {
                "id": str(q.id),
                "title": q.title,
                "complexity": q.complexity,
                "type": q.type,
                "options": q.options,
                "correct_answers": q.correct_answers,
                "max_score": q.max_score,
                "tags": q.tags,
            }
            for q in (exam.questions or [])
        ],
    }


@router.delete("/exams/{exam_id}")
def delete_exam(
    exam_id: UUID,
    db: Session = Depends(get_db),
    _: models.User = Depends(get_current_admin_user),
):
    """Delete an exam and all associated data."""
    exam = crud.get_exam_by_id(db, exam_id)
    if not exam:
        raise HTTPException(status_code=404, detail="Exam not found")
    
    # Delete all attempts and their answers
    attempts = db.query(models.ExamAttempt).filter(
        models.ExamAttempt.exam_id == exam_id
    ).all()
    
    for attempt in attempts:
        db.query(models.Answer).filter(
            models.Answer.attempt_id == attempt.id
        ).delete()
        db.delete(attempt)
    
    # Delete the exam
    db.delete(exam)
    db.commit()
    
    return {"status": "success", "message": "Exam deleted successfully"}


@router.get("/exams/{exam_id}/attempts")
def get_exam_attempts(
    exam_id: UUID,
    db: Session = Depends(get_db),
    _: models.User = Depends(get_current_admin_user),
):
    """Get all attempts for a specific exam."""
    # Verify exam exists
    exam = crud.get_exam_by_id(db, exam_id)
    if not exam:
        raise HTTPException(status_code=404, detail="Exam not found")
    
    # Get all attempts for this exam
    attempts = db.query(models.ExamAttempt).filter(
        models.ExamAttempt.exam_id == exam_id
    ).all()
    
    # Build response with student info and calculated scores
    result = []
    for attempt in attempts:
        student = db.query(models.User).filter(
            models.User.id == attempt.student_id
        ).first()
        
        # Get all answers for this attempt (needed for evaluation checking)
        answers = db.query(models.Answer).filter(
            models.Answer.attempt_id == attempt.id
        ).all()
        
        # Calculate final score with manual evaluations included
        manual_score_total = 0.0
        if attempt.end_time is not None:  # Only for submitted attempts
            for answer in answers:
                eval_record = crud.get_evaluation_by_answer(db, answer.id)
                if eval_record and eval_record.score_awarded is not None:
                    question = crud.get_question_by_id(db, answer.question_id)
                    if question and question.type in ("text", "image_upload"):
                        manual_score_total += float(eval_record.score_awarded)
        
        # Final score: auto-graded + manual scores (capped at total)
        final_score = min(attempt.score + manual_score_total, attempt.total_possible_score or 0) if attempt.score is not None else manual_score_total
        
        # Get evaluator info (admin who evaluated the exam)
        evaluator_email = None
        if answers:
            # Get the most recent evaluation to find who evaluated this attempt
            latest_eval = None
            for answer in answers:
                eval_record = crud.get_evaluation_by_answer(db, answer.id)
                if eval_record and (latest_eval is None or eval_record.updated_at > latest_eval.updated_at):
                    latest_eval = eval_record
            
            if latest_eval:
                evaluator = db.query(models.User).filter(models.User.id == latest_eval.evaluated_by).first()
                evaluator_email = evaluator.email if evaluator else "Unknown"
        
        attempt_data = {
            "id": str(attempt.id),
            "exam_id": str(attempt.exam_id),
            "student_id": str(attempt.student_id),
            "start_time": attempt.start_time.isoformat() if attempt.start_time else None,
            "end_time": attempt.end_time.isoformat() if attempt.end_time else None,
            "score": final_score,
            "total_possible_score": attempt.total_possible_score,
            "student": {
                "email": student.email if student else "Unknown",
                "full_name": student.full_name if (student and student.full_name) else (student.email if student else "Unknown"),
                "id": str(student.id) if student else None,
            },
            "is_submitted": attempt.end_time is not None,
            "evaluated_by": evaluator_email,
            "is_evaluated": evaluator_email is not None,
        }
        if attempt.end_time:
            attempt_data["submitted_at"] = attempt.end_time.isoformat()
        # Debug logging
        print(f"DEBUG: Student ID {attempt.student_id}, Email: {student.email if student else 'None'}, Full Name: {student.full_name if student else 'None'}")
        print(f"DEBUG: Attempt data student field: {attempt_data['student']}")
        result.append(attempt_data)
    
    return result


@router.get("/attempts/{attempt_id}/results")
def get_attempt_results_admin(
    attempt_id: UUID,
    db: Session = Depends(get_db),
    _: models.User = Depends(get_current_admin_user),
):
    """Get detailed results for a specific exam attempt (admin view)."""
    # Get the attempt
    attempt = db.query(models.ExamAttempt).filter(
        models.ExamAttempt.id == attempt_id
    ).first()
    
    if not attempt:
        raise HTTPException(status_code=404, detail="Attempt not found")
    
    # Get the student
    student = db.query(models.User).filter(
        models.User.id == attempt.student_id
    ).first()
    
    # Get the exam
    exam = crud.get_exam_by_id(db, attempt.exam_id)
    if not exam:
        raise HTTPException(status_code=404, detail="Exam not found")
    
    # Get student's answers
    answers = db.query(models.Answer).filter(
        models.Answer.attempt_id == attempt.id
    ).all()
    
    # Calculate final score with manual evaluations included
    manual_score_total = 0.0
    for answer in answers:
        eval_record = crud.get_evaluation_by_answer(db, answer.id)
        if eval_record and eval_record.score_awarded is not None:
            question = crud.get_question_by_id(db, answer.question_id)
            if question and question.type in ("text", "image_upload"):
                manual_score_total += float(eval_record.score_awarded)
    
    # Final score: auto-graded + manual scores (capped at total)
    final_score = min(attempt.score + manual_score_total, attempt.total_possible_score or 0) if attempt.score is not None else manual_score_total
    
    attempt_payload = schemas.ExamAttempt.model_validate(attempt).model_dump()
    attempt_payload["answers"] = [
        schemas.Answer.model_validate(a).model_dump() for a in answers
    ]
    attempt_payload["student_email"] = student.email if student else "Unknown"
    attempt_payload["score"] = final_score  # Override with calculated score
    
    # Get publisher email if published
    publisher_email = None
    if exam.published_by:
        publisher = db.query(models.User).filter(models.User.id == exam.published_by).first()
        publisher_email = publisher.email if publisher else "Unknown"
    
    # Manually build exam dict to avoid Pydantic validation issues with published_by
    exam_payload = {
        "id": str(exam.id),
        "title": exam.title,
        "start_time": exam.start_time.isoformat(),
        "end_time": exam.end_time.isoformat(),
        "duration_minutes": exam.duration_minutes,
        "is_published": exam.is_published,
        "published_by": publisher_email,
        "target_candidates": exam.target_candidates,
        "questions": [
            {
                "id": str(q.id),
                "title": q.title,
                "complexity": q.complexity,
                "type": q.type,
                "options": q.options,
                "correct_answers": q.correct_answers,
                "max_score": q.max_score,
                "tags": q.tags,
            }
            for q in (exam.questions or [])
        ],
    }
    
    return {"attempt": attempt_payload, "exam": exam_payload}


@router.post("/answers/{answer_id}/evaluate", response_model=schemas.Evaluation)
def evaluate_answer(
    answer_id: UUID,
    evaluation: schemas.EvaluationCreate,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_admin_user),
):
    """Evaluate a student's answer (mark as right/wrong, add comment)."""
    # Verify answer exists
    answer = db.query(models.Answer).filter(models.Answer.id == answer_id).first()
    if not answer:
        raise HTTPException(status_code=404, detail="Answer not found")
    
    # Ensure comment is max 100 characters
    if evaluation.comment and len(evaluation.comment) > 100:
        raise HTTPException(
            status_code=400,
            detail="Comment must be 100 characters or less"
        )
    
    eval_record = crud.create_or_update_evaluation(
        db, answer_id, current_user.id, evaluation
    )
    return eval_record


@router.get("/answers/{answer_id}/evaluation")
def get_answer_evaluation(
    answer_id: UUID,
    db: Session = Depends(get_db),
    _: models.User = Depends(get_current_admin_user),
):
    """Get evaluation for a specific answer."""
    evaluation = crud.get_evaluation_by_answer(db, answer_id)
    if not evaluation:
        return None
    return schemas.Evaluation.model_validate(evaluation).model_dump()


@router.get("/students/")
def list_all_students(
    db: Session = Depends(get_db),
    _: models.User = Depends(get_current_admin_user),
):
    """List all students (admin only)."""
    students = db.query(models.User).filter(models.User.role == "student").all()
    result = []
    for student in students:
        # Calculate overall performance percentage
        attempts = db.query(models.ExamAttempt).filter(
            models.ExamAttempt.student_id == student.id,
            models.ExamAttempt.end_time.isnot(None)  # Only completed exams
        ).all()
        
        overall_percentage = 0
        if attempts:
            total_percentage = sum([
                (attempt.score / attempt.total_possible_score * 100) if attempt.total_possible_score else 0
                for attempt in attempts
            ])
            overall_percentage = round(total_percentage / len(attempts))
        
        student_dict = {
            "id": str(student.id),
            "email": student.email,
            "full_name": student.full_name,
            "date_of_birth": student.date_of_birth.isoformat() if student.date_of_birth else None,
            "gender": student.gender,
            "exam_candidate": student.exam_candidate,
            "overall": overall_percentage,
        }
        result.append(student_dict)
    return result


@router.delete("/students/{student_id}")
def delete_student(
    student_id: UUID,
    db: Session = Depends(get_db),
    _: models.User = Depends(get_current_admin_user),
):
    """Delete a student and all their attempts and answers (admin only)."""
    # Verify student exists
    student = db.query(models.User).filter(
        models.User.id == student_id,
        models.User.role == "student"
    ).first()
    if not student:
        raise HTTPException(status_code=404, detail="Student not found")
    
    # Delete all attempts and their answers
    attempts = db.query(models.ExamAttempt).filter(
        models.ExamAttempt.student_id == student_id
    ).all()
    
    for attempt in attempts:
        # Delete evaluations for answers in this attempt
        answers = db.query(models.Answer).filter(
            models.Answer.attempt_id == attempt.id
        ).all()
        
        for answer in answers:
            db.query(models.Evaluation).filter(
                models.Evaluation.answer_id == answer.id
            ).delete()
            db.delete(answer)
        
        db.delete(attempt)
    
    # Delete the student
    db.delete(student)
    db.commit()
    
    return {"status": "success", "message": "Student deleted successfully with all their attempts and answers"}


@router.get("/exams/attempts/{attempt_id}")
def get_attempt_details(
    attempt_id: UUID,
    db: Session = Depends(get_db),
    _: models.User = Depends(get_current_admin_user),
):
    """Get detailed information about an exam attempt including all answers."""
    try:
        # Get the attempt
        attempt = db.query(models.ExamAttempt).filter(
            models.ExamAttempt.id == attempt_id
        ).first()
        
        if not attempt:
            raise HTTPException(status_code=404, detail="Attempt not found")
        
        # Get student and exam info
        student = db.query(models.User).filter(models.User.id == attempt.student_id).first()
        exam = db.query(models.Exam).filter(models.Exam.id == attempt.exam_id).first()
        
        # Build attempt response
        attempt_data = {
            "id": str(attempt.id),
            "student": {
                "id": str(student.id),
                "email": student.email,
                "full_name": student.full_name,
            } if student else None,
            "exam": {
                "id": str(exam.id),
                "title": exam.title,
            } if exam else None,
            "start_time": attempt.start_time.isoformat() if attempt.start_time else None,
            "end_time": attempt.end_time.isoformat() if attempt.end_time else None,
        }
        
        # Get all answers for this attempt
        answers = db.query(models.Answer).filter(
            models.Answer.attempt_id == attempt_id
        ).all()
        
        answers_data = []
        for answer in answers:
            question = db.query(models.Question).filter(
                models.Question.id == answer.question_id
            ).first()
            
            # Get evaluation data if exists
            evaluation = db.query(models.Evaluation).filter(
                models.Evaluation.answer_id == answer.id
            ).first()
            
            # Handle answer_data - it can be dict, list, string, or other types
            answer_text = None
            uploaded_file = None
            
            if isinstance(answer.answer_data, dict):
                # For dict types, try to extract text or name
                answer_text = answer.answer_data.get("text")
                uploaded_file = answer.answer_data.get("name")
            elif isinstance(answer.answer_data, (str, int, float, bool)):
                # For primitive types, use directly
                answer_text = str(answer.answer_data)
            elif isinstance(answer.answer_data, list):
                # For arrays, join them
                answer_text = ", ".join([str(x) for x in answer.answer_data])
            
            answer_dict = {
                "id": str(answer.id),
                "question_id": str(answer.question_id),
                "attempt_id": str(answer.attempt_id),
                "question": {
                    "id": str(question.id),
                    "title": question.title,
                    "type": question.type,
                    "correct_answers": question.correct_answers,
                } if question else None,
                "answer_text": answer_text,
                "uploaded_file": uploaded_file,
                "score": evaluation.score_awarded if evaluation else 0,
                "feedback": evaluation.comment if evaluation else "",
                "is_evaluated": evaluation is not None,
            }
            answers_data.append(answer_dict)
        
        return {
            "attempt": attempt_data,
            "answers": answers_data,
        }
    except HTTPException:
        raise
    except Exception as e:
        import traceback
        print(f"Error in get_attempt_details: {str(e)}")
        print(traceback.format_exc())
        raise HTTPException(status_code=500, detail=f"Internal server error: {str(e)}")


@router.post("/exams/attempts/{attempt_id}/evaluate")
def evaluate_exam_attempt(
    attempt_id: UUID,
    request_body: dict = Body(...),
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_admin_user),
):
    """Save evaluations for exam attempt answers."""
    # Get the attempt
    attempt = db.query(models.ExamAttempt).filter(
        models.ExamAttempt.id == attempt_id
    ).first()
    
    if not attempt:
        raise HTTPException(status_code=404, detail="Attempt not found")
    
    answers_data = request_body.get("answers", [])
    
    for answer_eval in answers_data:
        answer_id = UUID(str(answer_eval.get("answer_id")))
        score = answer_eval.get("score", 0)
        feedback = answer_eval.get("feedback", "")
        
        # Get or create evaluation record
        evaluation = db.query(models.Evaluation).filter(
            models.Evaluation.answer_id == answer_id
        ).first()
        
        if evaluation:
            # Update existing evaluation
            evaluation.score_awarded = score
            evaluation.comment = feedback
            evaluation.updated_at = datetime.now(timezone.utc)
        else:
            # Create new evaluation
            evaluation = models.Evaluation(
                answer_id=answer_id,
                evaluated_by=current_user.id,
                score_awarded=score,
                comment=feedback,
            )
            db.add(evaluation)
        
        db.commit()
    
    return {"status": "success", "message": "Evaluation saved successfully"}


@router.get("/evaluations/attempts/{attempt_id}")
def get_attempt_for_evaluation(
    attempt_id: UUID,
    db: Session = Depends(get_db),
    _: models.User = Depends(get_current_admin_user),
):
    """
    Get attempt details for evaluation page.
    Returns only text and image_upload type questions with their answers.
    """
    try:
        # Get the attempt
        attempt = db.query(models.ExamAttempt).filter(
            models.ExamAttempt.id == attempt_id
        ).first()
        
        if not attempt:
            raise HTTPException(status_code=404, detail="Attempt not found")
        
        # Get student info
        student = db.query(models.User).filter(
            models.User.id == attempt.student_id
        ).first()
        
        # Get exam info
        exam = db.query(models.Exam).filter(
            models.Exam.id == attempt.exam_id
        ).first()
        
        # Build response
        response = {
            "attempt_id": str(attempt.id),
            "student": {
                "id": str(student.id) if student else None,
                "email": student.email if student else None,
                "full_name": student.full_name if student else None,
            },
            "exam": {
                "id": str(exam.id) if exam else None,
                "title": exam.title if exam else None,
                "total_marks": sum([q.max_score for q in exam.questions]) if exam and exam.questions else 0,
            },
            "submitted_at": attempt.end_time.isoformat() if attempt.end_time else None,
            "answers": []
        }
        
        # Get all answers for this attempt
        answers = db.query(models.Answer).filter(
            models.Answer.attempt_id == attempt_id
        ).all()
        
        # Process each answer
        for answer in answers:
            question = db.query(models.Question).filter(
                models.Question.id == answer.question_id
            ).first()
            
            # Only include text and image_upload questions
            if not question or question.type not in ["text", "image_upload"]:
                continue
            
            # Get evaluation if exists
            evaluation = db.query(models.Evaluation).filter(
                models.Evaluation.answer_id == answer.id
            ).first()
            
            # Parse answer data
            answer_text = None
            uploaded_file_path = None
            uploaded_file_name = None
            
            if isinstance(answer.answer_data, dict):
                answer_text = answer.answer_data.get("text")
                uploaded_file_path = answer.answer_data.get("file_path")
                uploaded_file_name = answer.answer_data.get("name")
            elif isinstance(answer.answer_data, str):
                answer_text = answer.answer_data
            
            answer_dict = {
                "answer_id": str(answer.id),
                "question_id": str(question.id),
                "question_title": question.title,
                "question_type": question.type,
                "question_max_score": question.max_score,
                "answer_text": answer_text,
                "uploaded_file_path": uploaded_file_path,
                "uploaded_file_name": uploaded_file_name,
                "is_evaluated": evaluation is not None,
                "score_awarded": evaluation.score_awarded if evaluation else None,
                "comment": evaluation.comment if evaluation else None,
            }
            response["answers"].append(answer_dict)
        
        return response
        
    except HTTPException:
        raise
    except Exception as e:
        import traceback
        print(f"Error in get_attempt_for_evaluation: {str(e)}")
        print(traceback.format_exc())
        raise HTTPException(status_code=500, detail=f"Internal server error: {str(e)}")


@router.post("/evaluations/answers/{answer_id}/submit")
def submit_answer_evaluation(
    answer_id: UUID,
    evaluation_data: dict = Body(...),
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_admin_user),
):
    """
    Submit evaluation for a single answer.
    evaluation_data should have: {"is_correct": bool, "comment": str}
    """
    try:
        # Verify answer exists
        answer = db.query(models.Answer).filter(
            models.Answer.id == answer_id
        ).first()
        
        if not answer:
            raise HTTPException(status_code=404, detail="Answer not found")
        
        # Get the question to find max score
        question = db.query(models.Question).filter(
            models.Question.id == answer.question_id
        ).first()
        
        if not question:
            raise HTTPException(status_code=404, detail="Question not found")
        
        # Validate input
        is_correct = evaluation_data.get("is_correct")
        comment = evaluation_data.get("comment", "")
        
        if is_correct is None:
            raise HTTPException(status_code=400, detail="is_correct field is required")
        
        # Validate comment length
        if comment and len(comment) > 100:
            raise HTTPException(status_code=400, detail="Comment must be 100 characters or less")
        
        # Calculate score: full marks if correct, 0 if incorrect
        score_awarded = question.max_score if is_correct else 0
        
        # Get or create evaluation
        evaluation = db.query(models.Evaluation).filter(
            models.Evaluation.answer_id == answer_id
        ).first()
        
        if evaluation:
            evaluation.score_awarded = score_awarded
            evaluation.comment = comment
            evaluation.updated_at = datetime.now(timezone.utc)
        else:
            evaluation = models.Evaluation(
                answer_id=answer_id,
                evaluated_by=current_user.id,
                score_awarded=score_awarded,
                comment=comment,
            )
            db.add(evaluation)
        
        db.commit()
        
        return {
            "status": "success",
            "message": "Evaluation submitted successfully",
            "answer_id": str(answer_id),
            "score_awarded": score_awarded,
        }
        
    except HTTPException:
        raise
    except Exception as e:
        import traceback
        print(f"Error in submit_answer_evaluation: {str(e)}")
        print(traceback.format_exc())
        raise HTTPException(status_code=500, detail=f"Internal server error: {str(e)}")